import pandas as pd
import sys
import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side

def load_config():
    """Load configuration from config.json file"""
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
    
    # Use empty config if file doesn't exist
    if not os.path.exists(config_path):
        print(f"Warning: Configuration file not found: {config_path}")
        print("Using empty configuration.")
        return {"uppercase_words": []}
    
    try:
        with open(config_path, 'r') as config_file:
            config = json.load(config_file)
            
        # Validate that uppercase_words exists in the config
        if "uppercase_words" not in config:
            print("Warning: Configuration file is missing 'uppercase_words' key")
            print("Using empty uppercase_words configuration.")
            config["uppercase_words"] = []
            
        return config
    except json.JSONDecodeError:
        print(f"Warning: Invalid JSON format in configuration file: {config_path}")
        print("Using empty configuration.")
        return {"uppercase_words": []}
    except Exception as e:
        print(f"Warning: Error loading configuration file: {str(e)}")
        print("Using empty configuration.")
        return {"uppercase_words": []}

def format_header(header, uppercase_words):
    # Replace underscores with spaces
    words = header.replace('_', ' ').split()
    
    # Process each word
    formatted_words = []
    for word in words:
        # Keep certain words in all caps based on configuration
        if word.upper() in uppercase_words:
            formatted_words.append(word.upper())
        # All others are converted to title case
        else:
            formatted_words.append(word.title())
    
    return ' '.join(formatted_words)

def convert_csv_to_excel(csv_path):
    try:
        # Load configuration
        config = load_config()
        uppercase_words = config["uppercase_words"]
        
        # Read the CSV file
        df = pd.read_csv(csv_path)
        
        # Format column headers
        df.columns = [format_header(col, uppercase_words) for col in df.columns]
        
        # Generate output Excel file path
        excel_path = csv_path.rsplit('.', 1)[0] + '.xlsx'
        
        # Save to Excel
        df.to_excel(excel_path, index=False)
        
        # Load the workbook to apply formatting
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Freeze the top row
        ws.freeze_panes = 'A2'
        
        # Apply filter to the entire data range
        last_col = get_column_letter(ws.max_column)
        last_row = ws.max_row
        filter_range = f'A1:{last_col}{last_row}'
        ws.auto_filter.ref = filter_range
        
        # Left-align and remove borders from all cells
        left_alignment = Alignment(horizontal='left')

        no_border = Border(
            left=Side(style=None),
            right=Side(style=None),
            top=Side(style=None),
            bottom=Side(style=None)
        )
        
        # Apply formatting to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = left_alignment
                cell.border = no_border
        
        # Auto-fit column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            header_length = len(str(column[0].value))
            
            # Find the maximum length in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # If header is the longest, add extra padding for filter button
            if header_length >= max_length:
                max_length = header_length + 1  # Add extra padding for filter button
            
            # Set the column width (with a minimum of 10 and maximum of 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(excel_path)
        print(f"Successfully converted {csv_path} to {excel_path}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python excelify.py <path_to_csv_file>")
        sys.exit(1)
    
    csv_path = sys.argv[1]
    if not csv_path.lower().endswith('.csv'):
        print("Error: Input file must be a CSV file")
        sys.exit(1)
    
    convert_csv_to_excel(csv_path)
